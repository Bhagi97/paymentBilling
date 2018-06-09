from django.shortcuts import render
from django.template import loader
from django.http import HttpResponse
from django.http import JsonResponse
from django.core.files import File
from django.shortcuts import redirect
from django.contrib.auth.models import User, Permission
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

from interface.models import *

import pdb
import ast
import datetime
from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw
import os
import openpyxl

def login_request(request):
    context, perm_list = {}, {}

    pdb.set_trace()
    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        pdb.set_trace()
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request,user)
            return redirect('index')
        else:
            context['failure'] = True
            context['message'] = "Invalid Username/Password"
            template = loader.get_template('interface/login.html')

        return HttpResponse(template.render(context, request))

    if str(request.user)!='AnonymousUser':
        return redirect('index')

    template = loader.get_template('interface/login.html')
    return HttpResponse(template.render(context, request))



@login_required(login_url='login.html')
def logout_request(request):
    logout(request)
    template = loader.get_template('interface/login.html')
    return HttpResponse(template.render({}, request))


@login_required(login_url='login')
def index(request):
    if request.method == "POST":
        type = request.POST.get('type')
        if type=="Poster":
            kwargs = {"name" : request.POST.get('filename', ''),
                      "image" : request.FILES['posterToUpload'],
                      "uploader" : request.user,
                      }
            poster = Poster(**kwargs)
            poster.save()
        elif type=="Placement Shortlist":
            num = placementshortlist(request)
            for i in range(1,num+1):
                f = open(request.POST.get('name', 'Shortlist')+str(i)+'.jpg','rb')
                kwargs = {"name": request.POST.get('name', 'Shortlist')+str(i),
                          "uploader": request.user,
                          }
                poster = Poster(**kwargs)
                poster.image.save(request.POST.get('name', 'Shortlist')+str(i)+".jpg",File(f))
                poster.save()
                f.close()
                os.remove(request.POST.get('name', 'Shortlist')+str(i)+'.jpg')
    context = {}
    context['posters'] = Poster.objects.all()
    template = loader.get_template('interface/index.html')
    return HttpResponse(template.render(context, request))


def getposterdetails(request):
    data={}
    data['poster'] = Poster.objects.filter(id=request.GET['pid']).values()[0]
    data['poster']['created_at'] = datetime.datetime.strftime(data['poster']['created_at'],"%d/%m/%Y")
    data['success'] = 1
    return JsonResponse(data)

def updateposter(request):
    data={}
    p = Poster.objects.get(id=int(request.GET['id']))
    p.name = request.GET['data[filename]']
    p.show = True if request.GET.get('data[show]','off')=='on' else False
    p.save()
    data['success'] = 1
    return JsonResponse(data)


def deleteposter(request):
    data={}
    p = Poster.objects.filter(id=int(request.GET['id']))
    if len(p)>0:
        data['success'] = 1
        p.delete()
    else:
        data['success'] = 0
    return JsonResponse(data)

def placementshortlist(request):

    n = int(request.POST.get('colno', '2'))
    heading = request.POST.get('name', 'Shortlist')
    xl = request.FILES['xlfileToUpload']
    cl = request.FILES['companylogo']

    wb = openpyxl.load_workbook(xl)
    sheet = wb.active
    names=[]
    for i in range(1, sheet.max_row+1):
        name = sheet.cell(row=i, column=n).value
        if name != '':
            names.append(name.title())

    img = Image.open("Base.jpg")
    logo = Image.open(cl)
    draw = ImageDraw.Draw(img)
    font = ImageFont.truetype("arial.ttf",30)
    headerfont = ImageFont.truetype("arial.ttf",60)

    img, draw = setup_page(img,logo,draw,headerfont,heading)

    margin, offset =160, 400
    counter,pagecounter = 0,1
    for name in names:
        draw.text((margin, offset), str((pagecounter-1)*24+counter+1)+". "+name, font=font, fill="#000000")
        counter += 1
        offset += 50
        if counter == 24:
            img.save(heading + str(pagecounter) + '.jpg')
            img = Image.open("Base.jpg")
            draw = ImageDraw.Draw(img)
            img, draw = setup_page(img, logo, draw, headerfont, heading)
            margin, offset = 160, 400
            counter = 0
            pagecounter += 1
        elif counter == 12:
            margin, offset = 960, 400


    img.save(heading+ str(pagecounter) + '.jpg')

    return pagecounter


def setup_page(img,logo,draw,headerfont,heading):
    margin, offset = 960, 80
    c_w, c_h = logo.size
    if c_w >= 4 * c_h:
        af = 600.0 / c_w
    else:
        af = 150.0 / c_h
    c_w = int(round(af * c_w))
    c_h = int(round(af * c_h))
    logo = logo.resize((c_w, c_h), Image.ANTIALIAS)
    img.paste(logo, (margin - c_w // 2, offset))

    margin, offset = 960, 280
    draw.text((margin - draw.textsize(heading, font=headerfont)[0] // 2, offset), heading, font=headerfont,
              fill="#000000")

    return img,draw